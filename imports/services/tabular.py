"""
Read CSV or Excel uploads into a uniform list of dicts (v3.17.523).

The CSV importer previously did `csv.DictReader` inline in two places. Adding
spreadsheet support meant one reader both call sites share, so a .xlsx behaves
exactly like a .csv everywhere downstream — same headers, same row dicts, same
field mapping.

Scope note on .xls: openpyxl reads .xlsx/.xlsm only, and xlrd dropped legacy
.xls in 2.0. Rather than pull in an unmaintained pin, an .xls upload is refused
with a message telling the user to re-save as .xlsx — a clear error beats a
silent half-parse of a 1997 binary format.
"""
from __future__ import annotations

import csv
import io
import os

CSV_EXTENSIONS = ('.csv', '.txt', '.tsv')
EXCEL_EXTENSIONS = ('.xlsx', '.xlsm')
LEGACY_EXCEL_EXTENSIONS = ('.xls',)
SUPPORTED_EXTENSIONS = CSV_EXTENSIONS + EXCEL_EXTENSIONS


class TabularError(ValueError):
    """Raised with a message intended to be shown to the user."""


def _extension(filename: str) -> str:
    return os.path.splitext(filename or '')[1].lower()


def _clean(value) -> str:
    """Excel gives real types; the rest of the importer expects strings."""
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        # Excel stores every number as a float: quantity 12 arrives as 12.0,
        # and "12.0" would fail an IntegerField parse downstream.
        return str(int(value))
    return str(value).strip()


def _read_csv(file_obj, max_rows=None):
    file_obj.seek(0)
    content = file_obj.read()
    if isinstance(content, bytes):
        content = content.decode('utf-8-sig', errors='replace')
    reader = csv.DictReader(io.StringIO(content))
    headers = [h.strip() for h in (reader.fieldnames or []) if h is not None]
    rows = []
    for i, row in enumerate(reader):
        if max_rows is not None and i >= max_rows:
            break
        rows.append({(k.strip() if k else k): _clean(v) for k, v in row.items()})
    return headers, rows


def _read_excel(file_obj, max_rows=None):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:      # pragma: no cover - dependency is pinned
        raise TabularError(
            'Reading .xlsx files needs the openpyxl package. '
            'Run: pip install -r requirements.txt'
        ) from exc

    file_obj.seek(0)
    data = file_obj.read()
    try:
        # read_only keeps a large sheet from being loaded whole; data_only
        # takes the cached result of a formula rather than "=A1*B1".
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — surface a readable message
        raise TabularError(f'Could not read the spreadsheet: {exc}') from exc

    try:
        sheet = wb[wb.sheetnames[0]] if wb.sheetnames else None
        if sheet is None:
            raise TabularError('The spreadsheet has no sheets.')

        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], []

        headers = [_clean(h) for h in header_row]
        # Trailing empty columns are normal in hand-edited sheets.
        while headers and not headers[-1]:
            headers.pop()
        if not headers:
            raise TabularError('The first row of the sheet must contain column headers.')

        rows = []
        for i, raw in enumerate(rows_iter):
            if max_rows is not None and i >= max_rows:
                break
            values = list(raw) + [None] * (len(headers) - len(raw))
            row = {headers[c]: _clean(values[c]) for c in range(len(headers))}
            if any(row.values()):        # skip entirely blank rows
                rows.append(row)
        return headers, rows
    finally:
        wb.close()


def read_tabular(file_obj, filename: str = '', max_rows=None):
    """Return (headers, rows) for a CSV or Excel upload.

    `rows` is a list of dicts keyed by header, with every value a string, so
    callers cannot tell which format they were given.
    """
    ext = _extension(filename) or _extension(getattr(file_obj, 'name', ''))

    if ext in LEGACY_EXCEL_EXTENSIONS:
        raise TabularError(
            'Legacy .xls files are not supported. Open the file and use '
            '"Save As" → Excel Workbook (.xlsx), or export it as CSV.'
        )
    if ext in EXCEL_EXTENSIONS:
        return _read_excel(file_obj, max_rows=max_rows)
    if ext in CSV_EXTENSIONS or not ext:
        # No extension: assume CSV, which is what the importer accepted before.
        return _read_csv(file_obj, max_rows=max_rows)

    raise TabularError(
        f'Unsupported file type "{ext}". Upload a CSV or .xlsx spreadsheet.'
    )
